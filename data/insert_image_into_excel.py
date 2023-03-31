from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
import glob
from natsort import natsorted
import xlrd

# insert multiple images

# create a workbook and grab active worksheet
workbook = Workbook()
worksheet = workbook.active

workbook_open = xlrd.open_workbook('D:/Python/airflow_docker/data/ThanhPham.xls')
worksheet_open = workbook_open.sheet_by_index(0)

# images list
images = []
for filename in natsorted(glob.glob('D:/Python/airflow_docker/data/products/products/*.*')):
    images.append(filename)

# for filename in natsorted(glob.glob('D:/Python/airflow_docker/data/products/products/*.png')):
#     images.append(filename)

# for filename in natsorted(glob.glob('D:/Python/airflow_docker/data/products/products/*.jpeg')):
#     images.append(filename)

for r in range(worksheet_open.nrows):
    for c in range(worksheet_open.ncols):
        worksheet.cell(row=r+1, column=c+1).value = worksheet_open.cell(r, c).value
        worksheet.column_dimensions[get_column_letter(c+1)].width = 70

    # insert images
    for index, image in enumerate(images):
        if image==worksheet.cell(r+1, 7).value:
            img = Image(image)
            img.width = img.width*0.3
            img.height = img.height*0.3
            worksheet.row_dimensions[r+1].height = img.height
            worksheet.column_dimensions[get_column_letter(c+1)].width = img.width
            img.anchor = 'F'+str(r+1)
            worksheet.add_image(img)

# save workbook
workbook.save('ThanhPham02.xlsx')

