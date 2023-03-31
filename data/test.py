from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
import glob
import os
from natsort import natsorted
import xlrd
import xlwt


workbook = xlrd.open_workbook('D:/Python/airflow_docker/data/ThanhPham.xls')
worksheet = workbook.sheet_by_index(0)

new_workbook = xlwt.Workbook()
new_worksheet = new_workbook.add_sheet('Sheet1')

# images list
images = []
for filename in natsorted(glob.glob('D:/Python/airflow_docker/data/products/products/*.png')):
    images.append(filename)

for row in range(worksheet.nrows):
    for col in range(worksheet.ncols):
        new_worksheet.write(row, col, worksheet.cell(row, col).value)
        #worksheet.row_dimensions[row].height = 230
        #col_letter = get_column_letter(6)
        #worksheet.column_dimensions[col_letter].width = 40

    for index, image in enumerate(images):
        #print(image)
        if image==worksheet.cell(row, 6).value:
            #new_worksheet.insert_bitmap(Image(image), 5, index + 1)
            new_worksheet.add_image(Image(image), anchor='F'+str(index+1))

new_workbook.save('workbook.xls')